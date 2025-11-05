from openpyxl import load_workbook
from django.utils import timezone
from core.models import Kit
from datetime import date

SKIP_WORDS = ("setup", "travel")

def _scan_right(sheet, row_idx: int, start_col: int, max_checks: int = 20):
    """
    From (row_idx, start_col), move right until a cell that is:
      - not empty
      - and does NOT contain any SKIP_WORDS (case-insensitive)
    Returns (value_or_None, last_col_index_scanned)
    """
    col = start_col
    checks = 0
    while col <= sheet.max_column and checks < max_checks:
        v = sheet.cell(row=row_idx, column=col).value
        checks += 1
        if v:
            s = str(v).strip()
            if not any(w in s.lower() for w in SKIP_WORDS):
                return s, col  # found real value
        col += 1
    return None, col - 1  # not found

def update_kits_from_excel(file_path: str):
    wb = load_workbook(filename=file_path, data_only=True)
    sheet = wb.active

    # --- dynamic base columns by date ---
    base_date = date(2025, 11, 3)   # the day that maps to the first location column
    base_location_col = 477         # the column that corresponds to base_date
    days_since = (date.today() - base_date).days
    location_column_base = base_location_col + days_since  # shifts one col per day

    # NOTE: we will derive destination start per row after we find the location
    start_row = 10
    end_row = 19
    kit_num = 0
    now = timezone.now()

    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        kit_num += 1
        name = f"Kit {kit_num}"
        row_idx = row[0].row

        # 1) Scan for LOCATION starting from today's location base
        location, loc_col_found = _scan_right(sheet, row_idx, location_column_base)

        # 2) Scan for DESTINATION starting immediately after the location column
        destination_start_col = (loc_col_found or location_column_base) + 1
        destination, _ = _scan_right(sheet, row_idx, destination_start_col)

        # Update existing kit ONLY (no creation)
        kit = Kit.objects.filter(name__startswith=name.split('(')[0].strip()).first()
        if not kit:
            print(f"⚠️ No matching kit for {name}, skipping")
            continue

        kit.current_location = location or ""
        kit.destination_location = destination or ""
        kit.updated_at = now
        kit.save()

        print(f"{name}: current = {kit.current_location}, destination = {kit.destination_location}")

    print(f"Kits updated from {file_path}")
